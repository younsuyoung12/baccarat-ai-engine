# -*- coding: utf-8 -*-
# excel_logger.py
"""
Excel Logger for Baccarat Predictor AI Engine v12.1 (RULE-ONLY · STRICT)

역할:
- 날짜별 Excel 파일에 모든 슈/라운드 로그를 누적 저장
  (예: logs/excel/baccarat_2026-03-14.xlsx 에 해당 날짜의 모든 슈 기록)
- 헤더(EXCEL_COLUMNS) 관리
- app.py에서 전달한 row(dict)를 EXCEL_COLUMNS 순서에 맞춰 1행으로 기록
- /undo 시 같은 날짜 파일에서 해당 shoe_id 의 마지막 1행을 삭제

변경 요약 (2026-03-14)
----------------------------------------------------
1) GPT 전용 컬럼/설명 제거
   - GPT prediction / next GPT / prev_ai_* 계열 설명 제거
   - rule-only 엔진 기준 컬럼으로 재정의
2) STRICT 계약 정렬
   - CSV 폴백 제거
   - openpyxl 미설치 / 저장 실패 / 시트 손상 시 즉시 예외
   - 조용한 continue/pass 금지
3) 현재 app.py 계약 반영
   - bet_side / bet_unit / bet_reason / analysis / risk_tags / key_features
   - ai_total / ai_correct / ai_win_rate / ai_win_rate_pct
   - leader_* / road_hit_rates / future_* 요약
   - bet_tags_json / bet_metrics_json / future_scenarios_json / features_json 지원
4) is_correct 컬럼 유지
   - 명시적으로 전달된 값이 있으면 그것을 사용
   - 없으면 0으로 기록
----------------------------------------------------

주의:
- 이 모듈은 순수 로깅 전용이다.
- 로깅 실패는 상위 app.py 에서 롤백 처리해야 하므로, 이 모듈은 예외를 숨기지 않는다.
"""

from __future__ import annotations

import json
import os
import time
from datetime import datetime
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook, load_workbook  # type: ignore
except Exception:  # pragma: no cover
    Workbook = None
    load_workbook = None


LOG_BASE_DIR = "logs"
EXCEL_LOG_DIR = os.path.join(LOG_BASE_DIR, "excel")

os.makedirs(LOG_BASE_DIR, exist_ok=True)
os.makedirs(EXCEL_LOG_DIR, exist_ok=True)


EXCEL_COLUMNS: List[str] = [
    # 기본 메타
    "timestamp",
    "date",
    "shoe_id",
    "round_number",
    "winner",

    # 엔진/결과
    "ai_ok",
    "ai_error",
    "ai_engine",
    "bet_side",
    "bet_unit",
    "bet_reason",
    "entry_type",
    "analysis",

    # 누적 성능
    "ai_total",
    "ai_correct",
    "ai_win_rate",
    "ai_win_rate_pct",
    "ai_streak_win",
    "ai_streak_lose",
    "is_correct",

    # 핵심 feature
    "pattern_score",
    "pattern_reversal_signal",
    "flow_strength",
    "flow_chaos_risk",
    "flow_direction",

    # leader / confidence
    "leader_road",
    "leader_signal",
    "leader_confidence",
    "leader_trust_state",
    "road_hit_rates_json",

    # 운영/리스크
    "adaptive_chaos_limit",
    "reverse_bet_applied",
    "reverse_bet_original_side",

    # 미래 시나리오 요약
    "future_P_pattern_score",
    "future_P_flow_strength",
    "future_B_pattern_score",
    "future_B_flow_strength",

    # 태그 / 디버그
    "risk_tags",
    "key_features",
    "bet_tags_json",
    "bet_metrics_json",

    # 원본 JSON
    "future_scenarios_json",
    "features_json",
]


def new_shoe_id() -> str:
    """새 슈 ID 생성."""
    return datetime.now().strftime("shoe_%Y%m%d_%H%M%S")


def get_excel_path_for_date() -> str:
    """날짜별 Excel 로그 파일 경로 반환."""
    date_str = datetime.now().strftime("%Y-%m-%d")
    fname = f"baccarat_{date_str}.xlsx"
    return os.path.join(EXCEL_LOG_DIR, fname)


def _require_openpyxl() -> None:
    if Workbook is None or load_workbook is None:
        raise RuntimeError(
            "openpyxl is required for Excel logging but is not installed. "
            "Install openpyxl and retry."
        )


def _is_empty_row(values: List[Any]) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        return False
    return True


def _normalize_pb(v: Any) -> str:
    """P/B만 인정. 그 외(PASS/T/빈칸)는 ''."""
    if v is None:
        return ""
    s = str(v).strip().upper()
    if s in ("P", "PLAYER"):
        return "P"
    if s in ("B", "BANKER"):
        return "B"
    return ""


def _normalize_01(v: Any) -> Optional[int]:
    """0/1 정규화. 불가하면 None."""
    if v is None:
        return None
    if isinstance(v, bool):
        return 1 if v else 0
    try:
        s = str(v).strip().lower()
        if s in ("1", "true", "t", "yes", "y", "win", "success"):
            return 1
        if s in ("0", "false", "f", "no", "n", "lose", "fail", "loss"):
            return 0
        f = float(s)
        return 1 if f >= 0.5 else 0
    except Exception:
        return None


def _compute_is_correct(row: Dict[str, Any]) -> int:
    """
    is_correct 규칙:
    - 명시적으로 전달된 값(is_correct / prev_ai_correct / ai_correct / result_correct / correct)을 우선 사용
    - 없으면 0 기록
    """
    for k in ("is_correct", "prev_ai_correct", "result_correct", "correct"):
        if k in row:
            n01 = _normalize_01(row.get(k))
            if n01 is not None:
                return int(n01)

    # 현재 app.py row의 ai_correct 는 누적 정답 수이므로 단건 is_correct로 해석하면 안 된다.
    return 0


def _ensure_excel_header(ws) -> None:
    """
    엑셀 시트에 헤더가 없으면 1행에 생성/삽입/확장.
    """
    n = len(EXCEL_COLUMNS)
    first_row = [ws.cell(row=1, column=i + 1).value for i in range(n)]

    if _is_empty_row(first_row):
        for i, col_name in enumerate(EXCEL_COLUMNS, start=1):
            ws.cell(row=1, column=i, value=col_name)
        return

    if first_row == EXCEL_COLUMNS:
        return

    prefix_len = 0
    for i, col_name in enumerate(EXCEL_COLUMNS):
        v = first_row[i]
        if v == col_name:
            prefix_len += 1
            continue
        break

    if prefix_len >= 3:
        trailing = first_row[prefix_len:]
        trailing_empty = True
        for v in trailing:
            if v is None:
                continue
            if isinstance(v, str) and v.strip() == "":
                continue
            trailing_empty = False
            break

        if trailing_empty:
            for i, col_name in enumerate(EXCEL_COLUMNS, start=1):
                ws.cell(row=1, column=i, value=col_name)
            return

    ws.insert_rows(1)
    for i, col_name in enumerate(EXCEL_COLUMNS, start=1):
        ws.cell(row=1, column=i, value=col_name)


def _safe_json_dump(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, str):
        return v
    if isinstance(v, (dict, list, tuple, set)):
        try:
            return json.dumps(v, ensure_ascii=False)
        except Exception as e:
            raise RuntimeError(f"JSON serialization failed: {type(e).__name__}") from e
    return v


def _normalize_string_cell(v: Any, sep: str) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v
    if isinstance(v, list):
        parts: List[str] = []
        for x in v:
            if x is None:
                continue
            s = str(x).strip()
            if s:
                parts.append(s)
        return sep.join(parts)
    return str(v)


def _prepare_row(row: Dict[str, Any], shoe_id: Optional[str]) -> Dict[str, Any]:
    out = dict(row)

    out.setdefault("date", datetime.now().strftime("%Y-%m-%d"))
    out.setdefault("timestamp", datetime.now().isoformat(timespec="seconds"))

    if not out.get("shoe_id"):
        out["shoe_id"] = shoe_id or "unknown_shoe"

    # is_correct 강제 기록
    out["is_correct"] = int(_compute_is_correct(out))

    # 호환 입력 → 현재 컬럼 정규화
    if "road_hit_rates_json" not in out and "road_hit_rates" in out:
        out["road_hit_rates_json"] = out.get("road_hit_rates")

    if "bet_tags_json" not in out and "bet_tags" in out:
        out["bet_tags_json"] = out.get("bet_tags")

    if "bet_metrics_json" not in out and "bet_metrics" in out:
        out["bet_metrics_json"] = out.get("bet_metrics")

    if "future_scenarios_json" not in out and "future_scenarios" in out:
        out["future_scenarios_json"] = out.get("future_scenarios")

    if "features_json" not in out:
        if "features_raw" in out:
            out["features_json"] = out.get("features_raw")
        elif "features" in out:
            out["features_json"] = out.get("features")

    if "leader_trust_state" not in out and isinstance(out.get("bet_metrics"), dict):
        out["leader_trust_state"] = out["bet_metrics"].get("leader_trust_state")

    if "entry_type" not in out and isinstance(out.get("bet_metrics"), dict):
        out["entry_type"] = out["bet_metrics"].get("entry_type")

    return out


def _row_to_list(row: Dict[str, Any]) -> List[Any]:
    json_like_cols = {
        "road_hit_rates_json",
        "bet_tags_json",
        "bet_metrics_json",
        "future_scenarios_json",
        "features_json",
    }

    out: List[Any] = []
    for col in EXCEL_COLUMNS:
        v = row.get(col)

        if col == "risk_tags":
            v = _normalize_string_cell(v, ",")
        elif col == "key_features":
            v = _normalize_string_cell(v, "|")
        elif col in json_like_cols:
            v = _safe_json_dump(v)

        out.append(v)
    return out


def _append_excel_row_with_retry(
    excel_path: str,
    row_values: List[Any],
    retries: int = 3,
    delay: float = 0.30,
) -> None:
    _require_openpyxl()

    last_exc: Optional[Exception] = None
    for attempt in range(retries):
        try:
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
            else:
                wb = Workbook()

            ws = wb.active
            _ensure_excel_header(ws)
            ws.append(row_values)
            wb.save(excel_path)
            return

        except PermissionError as e:
            last_exc = e
            if attempt < retries - 1:
                time.sleep(delay)
                continue
            raise RuntimeError(
                f"Excel file is locked and could not be saved after {retries} attempts: {excel_path}"
            ) from e

        except Exception as e:
            last_exc = e
            raise RuntimeError(f"Failed to append Excel row: {excel_path}") from e

    if last_exc is not None:
        raise RuntimeError(f"Failed to append Excel row: {excel_path}") from last_exc


def append_round_log_to_excel(row: Dict[str, Any], shoe_id: Optional[str]) -> None:
    """한 라운드 로그를 날짜별 Excel에 추가한다."""
    if not isinstance(row, dict):
        raise TypeError(f"row must be dict, got {type(row).__name__}")

    excel_path = get_excel_path_for_date()
    prepared_row = _prepare_row(row, shoe_id)
    row_list = _row_to_list(prepared_row)
    _append_excel_row_with_retry(excel_path, row_list)


def remove_last_round_log_for_shoe(shoe_id: str) -> None:
    """
    현재 날짜 엑셀 로그에서 주어진 shoe_id 의 마지막 1행을 삭제한다.
    - 파일이 없거나 해당 shoe_id 행이 없으면 그대로 반환
    - 예외는 숨기지 않는다
    """
    if not isinstance(shoe_id, str) or not shoe_id.strip():
        raise ValueError("shoe_id must be non-empty string")

    _require_openpyxl()

    excel_path = get_excel_path_for_date()
    if not os.path.exists(excel_path):
        return

    wb = load_workbook(excel_path)
    ws = wb.active

    _ensure_excel_header(ws)

    try:
        shoe_col_idx = EXCEL_COLUMNS.index("shoe_id") + 1
    except ValueError as e:
        raise RuntimeError("EXCEL_COLUMNS missing required column: shoe_id") from e

    target_row: Optional[int] = None
    for row_idx in range(ws.max_row, 1, -1):
        cell_val = ws.cell(row=row_idx, column=shoe_col_idx).value
        if cell_val == shoe_id:
            target_row = row_idx
            break

    if target_row is None:
        return

    ws.delete_rows(target_row, 1)
    wb.save(excel_path)